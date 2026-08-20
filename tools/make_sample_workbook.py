"""
Build deliverable/sample_inventory.xlsx.

A safe stand-in for the real inventory sheet, shipped next to the
installer so anyone can try the app, or test a change, without a
copy of live company data.

It mirrors the structure of the production workbook rather than
inventing a tidier one: the same header row, the same column
formats and widths, and the same defined table over the used
range. Item numbers and product names are invented; the figures
are made up.

Two things here are deliberately harder than the current
production file, because they are the cases most likely to break:

  - The diff column holds formulas. The app must never write to
    that column, so a round trip through upload and download has
    to leave them intact.
  - Two item numbers (PR40105, RAW0003) do not match the pattern
    in app/validators.py. That pattern only governs items added
    by hand in the app; anything already in the workbook is
    imported as it stands, and the real sheet does contain rows
    like these.

Run from the project root:

    .venv\\Scripts\\python.exe tools\\make_sample_workbook.py
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


PROJECT_ROOT = Path(__file__).resolve().parent.parent

OUTPUT_WORKBOOK = (
    PROJECT_ROOT
    / "deliverable"
    / "sample_inventory.xlsx"
)

HEADERS = [
    "Item number",
    "Product name",
    "Physical inventory",
    "counted",
    "diff",
]

# Matches the production sheet, so the sample looks like the real
# thing when it is opened.
COLUMN_WIDTHS = {
    "A": 15.0,
    "B": 64.0,
    "C": 22.0,
    "D": 21.0,
    "E": 13.0,
}

TEXT_FORMAT = "@"

INVENTORY_FORMAT = "#,##0.00"

# Item number, product name, physical inventory.
ITEMS = [
    ("ALL8109", "FILTERS PAPER WIDE BASE (S)", 40.20),
    ("FLA9200", "SAMPLE SPICE NATURAL & ARTIFICIAL", 1.27),
    ("FLA9205", "SAMPLE CRUNCH NATURAL & ARTIFICIAL", 779.92),
    ("FLA9289-2", "SAMPLE BLUES NATURAL & ARTIFICIAL (BC)", 88.50),
    ("FLA9292-1", "SAMPLE PECAN NATURAL & ARTIFICIAL", 73.79),
    ("FLA9302", "SAMPLE TOFFEE (S)", 38.90),
    ("PKG0628", "SAMPLE CARTON 12 X 340 G", 2140.00),
    ("PKG1422-2", "SAMPLE CARTON 6 X 1 KG (S)", 615.50),
    ("PKG1730", "SAMPLE VALVE 8 MM WHITE", 18400.00),
    ("PKG2044", "SAMPLE TIN TIE BAG 340 G KRAFT", 7320.75),
    ("PKG2810", "SAMPLE LABEL ROLL 4 X 6 THERMAL", 96.00),
    ("PKG3155", "SAMPLE SHRINK FILM 16 IN CLEAR", 1245.30),
    ("PKG4001-1", "SAMPLE PALLET SHEET 40 X 48", 310.00),
    ("PKG4610", "SAMPLE CASE TAPE 48 MM CLEAR", 288.00),
    ("PKG5233", "SAMPLE STICK PACK FILM 2.5 G", 43610.88),
    ("PKG6120", "SAMPLE K CUP LID FOIL (2026)", 22750.00),
    ("PKG7044-3", "SAMPLE POUCH 5 LB FOIL LINED", 1902.40),
    ("PKG8115", "SAMPLE CORRUGATE TRAY 24 CT", 1560.00),
    ("PKG9965", "SAMPLE WHITE 16 OZ 364 MM FILM (S)", 12535.81),
    ("PKG9966", "SAMPLE RETAIL FILM 7.5 OZ (2026)", 3600.00),
    # These two prefixes are outside the manual-entry pattern on
    # purpose. See the module docstring.
    ("PR40105", "SAMPLE FLAGSHIP 24 X 2.50 OZ (S)", 1.00),
    ("RAW0003", "SAMPLE ICING SUGAR SIX X 1 LB", 537.88),
]

TABLE_NAME = "Table"

TABLE_STYLE = "TableStyleMedium2"


def build_worksheet(worksheet) -> None:
    worksheet.append(HEADERS)

    for row_offset, (
        item_number,
        product_name,
        physical_inventory,
    ) in enumerate(ITEMS):
        row = row_offset + 2

        worksheet.cell(
            row=row,
            column=1,
        ).value = item_number

        worksheet.cell(
            row=row,
            column=2,
        ).value = product_name

        inventory_cell = worksheet.cell(
            row=row,
            column=3,
        )

        inventory_cell.value = physical_inventory
        inventory_cell.number_format = INVENTORY_FORMAT

        # counted is left empty. The app fills it in on export,
        # and a workbook arriving with values there would hide a
        # bug in that step.

        worksheet.cell(
            row=row,
            column=5,
        ).value = f"=C{row}-D{row}"

    for column in (1, 2):
        for row in range(2, len(ITEMS) + 2):
            worksheet.cell(
                row=row,
                column=column,
            ).number_format = TEXT_FORMAT

    for column_letter, width in COLUMN_WIDTHS.items():
        worksheet.column_dimensions[
            column_letter
        ].width = width


def add_table(worksheet) -> None:
    """
    The production sheet has a defined table over its used range,
    and manually added items land below it. Keeping the table
    here means that case gets exercised by anyone testing with
    this file.
    """
    last_column = get_column_letter(len(HEADERS))

    table = Table(
        displayName=TABLE_NAME,
        ref=f"A1:{last_column}{len(ITEMS) + 1}",
    )

    table.tableStyleInfo = TableStyleInfo(
        name=TABLE_STYLE,
        showRowStripes=True,
    )

    worksheet.add_table(table)


def main() -> None:
    OUTPUT_WORKBOOK.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Sheet1"

    build_worksheet(worksheet)
    add_table(worksheet)

    workbook.save(OUTPUT_WORKBOOK)
    workbook.close()

    print(
        f"Wrote {OUTPUT_WORKBOOK} "
        f"({len(ITEMS)} items)"
    )


if __name__ == "__main__":
    main()

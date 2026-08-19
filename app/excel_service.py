from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from app.config import HEADER_ALIASES, WORKBOOK_PATH
from app.database import db_connect, get_item_count_summary


def normalize_header(value: Any) -> str:
    return (
        str(value or "")
        .strip()
        .lower()
        .replace("_", " ")
    )


def find_header_row_and_columns(worksheet):
    maximum_header_row = min(
        worksheet.max_row,
        20,
    )

    for row_number in range(
        1,
        maximum_header_row + 1,
    ):
        headers = {
            normalize_header(
                worksheet.cell(
                    row=row_number,
                    column=column_number,
                ).value
            ): column_number
            for column_number in range(
                1,
                worksheet.max_column + 1,
            )
            if worksheet.cell(
                row=row_number,
                column=column_number,
            ).value is not None
        }

        found_columns: dict[str, int] = {}

        for logical_name, aliases in HEADER_ALIASES.items():
            for alias in aliases:
                if alias in headers:
                    found_columns[logical_name] = headers[alias]
                    break

        if (
            "item_number" in found_columns
            and "counted" in found_columns
        ):
            return row_number, found_columns

    raise ValueError(
        "The Excel file must contain an Item number column "
        "and a counted column."
    )


def ensure_added_manually_column(
    worksheet,
    header_row: int,
) -> int:
    for column_number in range(
        1,
        worksheet.max_column + 1,
    ):
        header = normalize_header(
            worksheet.cell(
                row=header_row,
                column=column_number,
            ).value
        )

        if header == "added manually":
            return column_number

    new_column = worksheet.max_column + 1

    worksheet.cell(
        row=header_row,
        column=new_column,
    ).value = "Added Manually"

    return new_column


def validate_workbook_file(workbook_path) -> None:
    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=False,
    )

    try:
        worksheet = workbook.active
        find_header_row_and_columns(worksheet)

    finally:
        workbook.close()


def import_excel_to_db() -> None:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(
            "The source Excel file does not exist."
        )

    workbook = load_workbook(
        WORKBOOK_PATH,
        data_only=False,
    )

    try:
        worksheet = workbook.active

        header_row, columns = (
            find_header_row_and_columns(worksheet)
        )

        item_column = columns["item_number"]
        product_column = columns.get("product_name")
        inventory_column = columns.get(
            "physical_inventory"
        )

        imported_count = 0

        with db_connect() as connection:
            for row_number in range(
                header_row + 1,
                worksheet.max_row + 1,
            ):
                item_number = str(
                    worksheet.cell(
                        row=row_number,
                        column=item_column,
                    ).value
                    or ""
                ).strip().upper()

                if not item_number:
                    continue

                product_name = ""

                if product_column:
                    product_name = str(
                        worksheet.cell(
                            row=row_number,
                            column=product_column,
                        ).value
                        or ""
                    ).strip()

                physical_inventory = ""

                if inventory_column:
                    physical_inventory = str(
                        worksheet.cell(
                            row=row_number,
                            column=inventory_column,
                        ).value
                        or ""
                    ).strip()

                connection.execute(
                    """
                    INSERT OR REPLACE INTO items (
                        item_number,
                        product_name,
                        physical_inventory,
                        row_number,
                        added_manually
                    )
                    VALUES (?, ?, ?, ?, '')
                    """,
                    (
                        item_number,
                        product_name,
                        physical_inventory,
                        row_number,
                    ),
                )

                imported_count += 1

        if imported_count == 0:
            raise ValueError(
                "No inventory items were found "
                "in the uploaded Excel file."
            )

    finally:
        workbook.close()


def build_export_workbook() -> BytesIO:
    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(
            "No workbook has been uploaded."
        )

    workbook = load_workbook(
        WORKBOOK_PATH,
        data_only=False,
    )

    worksheet = workbook.active

    header_row, columns = (
        find_header_row_and_columns(worksheet)
    )

    item_column = columns["item_number"]
    counted_column = columns["counted"]
    product_column = columns.get("product_name")

    added_manually_column = (
        ensure_added_manually_column(
            worksheet,
            header_row,
        )
    )

    with db_connect() as connection:
        items = connection.execute(
            """
            SELECT *
            FROM items
            ORDER BY item_number
            """
        ).fetchall()

    for item in items:
        item_number = item["item_number"]

        total, entry_count = (
            get_item_count_summary(item_number)
        )

        exported_count = (
            total
            if entry_count > 0
            else None
        )

        if item["added_manually"] == "Yes":
            new_row = worksheet.max_row + 1

            worksheet.cell(
                row=new_row,
                column=item_column,
            ).value = item_number

            if product_column:
                worksheet.cell(
                    row=new_row,
                    column=product_column,
                ).value = item["product_name"]

            worksheet.cell(
                row=new_row,
                column=counted_column,
            ).value = exported_count

            worksheet.cell(
                row=new_row,
                column=added_manually_column,
            ).value = "Yes"

        else:
            row_number = item["row_number"]

            if row_number:
                worksheet.cell(
                    row=row_number,
                    column=counted_column,
                ).value = exported_count

    output = BytesIO()

    workbook.save(output)
    workbook.close()

    output.seek(0)

    return output